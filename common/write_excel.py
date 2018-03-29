# coding:utf-8
import openpyxl
from openpyxl import load_workbook
import xlrd
import os
from log import Log
def copy_excel(path1,path2):
    wb=openpyxl.Workbook()
    wb.save(path2)
    #读取数据
    wb1=load_workbook(path1)
    wb2=load_workbook(path2)
    sheets1=wb1.sheetnames
    sheets2=wb2.sheetnames
    sheet1=wb1[sheets1[0]]
    sheet2=wb2[sheets2[0]]
    print sheet1,sheet2
    #获取行和列
    max_row=sheet1.max_row
    max_column=sheet1.max_column
    # print max_row,max_column

    #复制数据
    try:
        for i in range(max_row):
            if i<1:
                sheet2.cell(1,1).value = "id"
            else:
                sheet2.cell(i+1,1).value=i
            for j in range(max_column):
                value=sheet1.cell(i+1,j+1).value
                sheet2.cell(i+1,j+2).value=value
        wb2.save(path2)
        log=Log()
        log.info("copy excel is successful")


    except Exception as msg:
        print(msg)

class UpdataData():
    def __init__(self,file_path,sheet_index=0):
        self.file_path=file_path
        self.wb=load_workbook(file_path)
        self.sheets=self.wb.sheetnames
        self.sheet=self.wb[self.sheets[sheet_index]]
        self.log=Log()

    def updata_value(self,i,j,value):
        #给单元格赋值
        try:
            self.sheet.cell(i,j).value=value
            self.wb.save(self.file_path)
            self.log.info("(%s,%s)updata successful!!"%(i,j))
        except Exception as msg:
            print msg
#
if __name__=="__main__":
    data_path=os.path.join(os.path.dirname(os.path.dirname(os.path.realpath("__file__"))),"data\\kuaidi.xlsx")
    copy_path = os.path.join(os.path.dirname(os.path.dirname(os.path.realpath("__file__"))), "data\\kuaidi_copy.xlsx")
    copy_excel(data_path,copy_path)
    updata=UpdataData(copy_path)
    updata.updata_value(1,1,"")
